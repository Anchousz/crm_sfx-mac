# -*- coding: utf-8 -*-
"""Импорт прайса из Excel и выгрузка сметы."""
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import paths

FONT = "Arial"
UNIT_HINTS = (("жидкость", "л"), ("километраж", "км"), ("кв.м", "кв.м"),
              ("погрузка", "ч"), ("разгрузка", "ч"))


def parse_date(label):
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(label).strip(), fmt)
        except ValueError:
            pass
    return None


def guess_unit(name):
    low = name.lower()
    for token, unit in UNIT_HINTS:
        if token in low:
            return unit
    return "шт"


def import_prices(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = []
    current = None
    prev_blank = True
    first = True
    for row in ws.iter_rows(min_row=1, max_col=2):
        name = row[0].value
        price = row[1].value
        if first:
            first = False
            continue
        if name is None or not str(name).strip():
            prev_blank = True
            continue
        name = str(name).strip()
        if prev_blank:
            current = (name, [])
            result.append(current)
        elif current is not None:
            try:
                p = float(price)
            except (TypeError, ValueError):
                p = 0.0
            current[1].append((name, p, guess_unit(name)))
        prev_blank = False
    return [(c, items) for c, items in result if items]


def export_estimate(path, project, company=None):
    company = company or {}
    days = project["days"]
    tax = project.get("tax") or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"

    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, size=10, bold=True)
    big = Font(name=FONT, size=11, bold=True)
    muted = Font(name=FONT, size=9, color="808080", italic=True)
    thin = Side(style="thin", color="BFBFBF")
    top_line = Border(top=thin)
    head_fill = PatternFill("solid", fgColor="F2F2F2")
    money = "#,##0"
    right = Alignment(horizontal="right")

    widths = {"A": 16, "B": 46, "C": 11, "D": 9, "E": 7, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    r = 1
    if company.get("company_name"):
        ws.cell(r, 1, company["company_name"]).font = Font(name=FONT, size=13, bold=True)
        r += 1
        contacts = " · ".join(x for x in (company.get("company_phone"), company.get("company_email")) if x)
        if contacts:
            ws.cell(r, 1, contacts).font = muted
            r += 1
        if company.get("company_extra"):
            ws.cell(r, 1, company["company_extra"]).font = muted
            r += 1
        logo = paths.logo_path()
        if logo:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(logo)
                scale = 52.0 / img.height
                img.height = int(img.height * scale)
                img.width = int(img.width * scale)
                ws.add_image(img, "F1")
            except Exception:
                pass
        r += 1

    title = project["name"] + (f" — {project['client']}" if project.get("client") else "")
    ws.cell(r, 1, title).font = Font(name=FONT, size=12, bold=True)
    sub = f"Смета от {datetime.now():%d.%m.%Y} · статус: {project.get('status', 'предварительная')}"
    ws.cell(r + 1, 1, sub).font = muted
    r += 3

    headers = ["Дата", "Наименование", "Цена", "Кол-во", "Ед.", "Стоимость"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = Border(bottom=thin)
    r += 2

    subtotal_cells = []
    ot_cells = []
    for day in days:
        lines = day["lines"]
        if not lines:
            continue
        first_row = r
        label = day["label"]
        d = parse_date(label)
        a = ws.cell(first_row, 1, d if d else label)
        a.font = bold
        if d:
            a.number_format = "DD.MM.YYYY"
        if day.get("note"):
            ws.cell(first_row, 2, day["note"]).font = muted
            r += 1
        for ln in lines:
            is_ot = bool(ln.get("ot"))
            name = ln["name"] + (" — переработка" if is_ot else "")
            ws.cell(r, 2, name).font = base
            c3 = ws.cell(r, 3, ln["price"])
            c3.font = base
            c3.number_format = money
            c4 = ws.cell(r, 4, ln["qty"])
            c4.font = base
            c4.number_format = "0.##"
            ws.cell(r, 5, "ч" if is_ot else ln.get("unit", "шт")).font = base
            c6 = ws.cell(r, 6, f"=C{r}*D{r}")
            c6.font = base
            c6.number_format = money
            if is_ot:
                ot_cells.append(f"F{r}")
            r += 1
        sub_row_first = first_row + (1 if day.get("note") else 0)
        sub = ws.cell(r, 6, f"=SUM(F{sub_row_first}:F{r - 1})")
        sub.font = bold
        sub.number_format = money
        sub.border = top_line
        sub.alignment = right
        subtotal_cells.append(f"F{r}")
        r += 2

    if not subtotal_cells:
        wb.save(path)
        return

    def money_cell(row, formula, font):
        cell = ws.cell(row, 6, formula)
        cell.font = font
        cell.number_format = money
        cell.alignment = right
        return cell

    subs = "+".join(subtotal_cells)
    ot = "+".join(ot_cells) if ot_cells else "0"
    r += 1
    ws.cell(r, 2, "Позиции, без учёта налога и переработок").font = bold
    money_cell(r, f"=({subs})-F{r + 1}", bold)
    ws.cell(r + 1, 2, "Переработки").font = bold
    money_cell(r + 1, f"={ot}", bold)
    ws.cell(r + 2, 2, "Итого до налога").font = bold
    money_cell(r + 2, f"=F{r}+F{r + 1}", bold)
    ws.cell(r + 3, 2, "Налог").font = bold
    tax_cell = ws.cell(r + 3, 4, tax / 100.0)
    tax_cell.font = Font(name=FONT, size=10, bold=True, color="0000FF")
    tax_cell.number_format = "0.#%"
    money_cell(r + 3, f"=F{r + 2}*D{r + 3}", bold)
    ws.cell(r + 4, 2, "ИТОГО").font = big
    total = money_cell(r + 4, f"=F{r + 2}+F{r + 3}", big)
    total.border = Border(top=Side(style="double", color="404040"))

    wb.save(path)